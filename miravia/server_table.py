from flask import Flask, render_template, request, jsonify, send_file, make_response
import psycopg2
import json
from datetime import datetime
import html, re, json
import openpyxl
from openpyxl import load_workbook
import io
from io import BytesIO


app = Flask(__name__)
app.config['DATABASE_URI'] = 'postgresql://xxx:xxxx@192.168.xxx.xxx/postgres'

class User:
    def __init__(self, assignment_date, lead_number, source, url, product_service, is_local_spanish_seller, vat, cluster, l1, l2, company_name, revenue, product_count, product_desc_or_at_least_2_product_images, amount_phone_number, telephone, telephone1, amount_emails, email, email1, feedback_final, qualification_feedback, resolution_date, leads_id, prio,pic=''):
        self.pic = pic
        self.assignment_date = assignment_date
        self.lead_number = lead_number
        self.source = source
        self.url = url
        self.product_service = product_service
        self.is_local_spanish_seller = is_local_spanish_seller
        self.vat = vat
        self.cluster = cluster
        self.l1 = l1
        self.l2 = l2
        self.company_name = company_name
        self.revenue = revenue
        self.product_count = product_count
        self.product_desc_or_at_least_2_product_images = product_desc_or_at_least_2_product_images
        self.amount_phone_number = amount_phone_number
        self.telephone = telephone
        self.telephone1 = telephone1
        self.amount_emails = amount_emails
        self.email = email
        self.email1 = email1
        self.feedback_final = feedback_final
        self.qualification_feedback = qualification_feedback    
        self.resolution_date = resolution_date
        self.leads_id = leads_id
        self.prio = prio



def file_name_exists(file_name):
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    select_query = """
        SELECT file_name
        FROM miravia.miravia_table
        WHERE file_name = %s
        LIMIT 1
    """

    cursor.execute(select_query, (file_name,))
    result = cursor.fetchone()

    cursor.close()
    conn.close()

    return result is not None

@app.route('/check_file', methods=['POST'])
def check_file():
    file = request.files['file']
    if file_name_exists(file.filename):
        return jsonify({'exists': True})
    else:
        return jsonify({'exists': False})


@app.route('/import', methods=['POST'])
def import_data():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    file = request.files['file']
    start_column = "Assignment Date"
    end_column = "Prio"

    try:
        wb = openpyxl.load_workbook(file)
        
        column_mapping = {
            'Assignment Date': 'assignment_date',
            'LEAD NUMBER': 'lead_number',
            'Source': 'source',
            'LINK/URL': 'url',
            'PRODUCT/SERVICE': 'product_service',
            'LOCAL SPANISH SELLER': 'local_spanish_seller',
            'VAT': 'vat',
            'CLUSTER': 'cluster',
            'L1': 'l1',
            'L2': 'l2',
            'COMPANY NAME': 'company_name',
            'REVENUE': 'revenue',
            'PRODUCT COUNT': 'product_count',
            'PRODUCT DESCRIPTION OR AT LEAST 2 PRODUCT IMAGES': 'product_desc_or_at_least_2_product_images',
            'AMOUNT OF PHONE NUMBERS?': 'amount_phone_number',
            'TELEPHONE': 'telephone',
            'TELEPHONE1': 'telephone1',
            'AMOUNT OF EMAILS': 'amount_emails',
            'EMAIL': 'email',
            'EMAIL1': 'email1',
            'Feedback Final': 'feedback_final',
            'QUALIFICATION FEEDBACK': 'qualification_feedback',
            'Resolution Date': 'resolution_date',
            'Leads ID': 'leads_id',
            'Prio': 'prio',
            'Lead ID': 'lead_id',
        }

        user_data_list = []
        # Check if the "High Priority" sheet exists
        if 'High Priority' in wb.sheetnames:
            print('High Priority',len(user_data_list))
            ws = wb['High Priority']
            if ws.sheet_state == 'hidden':
                print('High Priority is hidden')
                pass
            else:
                high_priority_column_names = [cell.value for cell in ws[1]]
                high_priority_min_row = 2
                if 'Assignment Date' not in high_priority_column_names:
                    high_priority_column_names = [cell.value for cell in ws[2]]
                    high_priority_min_row = 3

                high_priority_start_index = high_priority_column_names.index(start_column)
                high_priority_end_index = high_priority_column_names.index(end_column)

                for row in ws.iter_rows(min_row=high_priority_min_row):
                    # Check if any data is present in the row
                    if any(row[i].value for i in range(high_priority_start_index, high_priority_end_index + 1)):
                        high_priority_data = {
                            column_mapping.get(high_priority_column_names[i], high_priority_column_names[i]): str(row[i].value) if row[i].value is not None else ''
                            for i in range(high_priority_start_index, high_priority_end_index + 1)
                        }
                        high_priority_data['import_date'] = datetime.now().strftime("%Y-%m-%d")  # Add the import_date value
                        user_data_list.append((
                            re.sub('\s.*', '', str(high_priority_data['assignment_date'])),
                            re.sub('\.0$|_\d{8}', '', str(high_priority_data['lead_number']))+'_%s'%datetime.now().strftime("%Y%m%d"),
                            high_priority_data['source'],
                            high_priority_data['url'],
                            high_priority_data['vat'],
                            re.sub('\.0$', '', high_priority_data['amount_phone_number'] or ''),
                            re.sub('\.0$', '', high_priority_data['amount_emails'] or ''),
                            file.filename,
                            high_priority_data['import_date'],
                            high_priority_data['import_date']+'_high_prio',
                            high_priority_data['local_spanish_seller'],
                            high_priority_data['cluster'],
                            high_priority_data['company_name'],
                            high_priority_data['product_count'],
                            high_priority_data['revenue'],
                            high_priority_data['telephone'],
                            high_priority_data['telephone1'],
                            high_priority_data['email'],
                            high_priority_data['email1'],
                            high_priority_data['lead_id'],
                        ))

        # Check if the "Middle High" sheet exists
        if 'Middle High' in wb.sheetnames:
            print('Middle High',len(user_data_list))
            ws = wb['Middle High']
            if ws.sheet_state == 'hidden':
                print('Middle High is hidden')
                pass
            else:
                middle_high_column_names = [cell.value for cell in ws[1]]
                middle_high_min_row = 2
                if 'Assignment Date' not in middle_high_column_names:
                    middle_high_column_names = [cell.value for cell in ws[2]]
                    middle_high_min_row = 3

                middle_high_start_index = middle_high_column_names.index(start_column)
                middle_high_end_index = middle_high_column_names.index(end_column)

                for row in ws.iter_rows(min_row=middle_high_min_row):
                    # Check if any data is present in the row
                    if any(row[i].value for i in range(middle_high_start_index, middle_high_end_index + 1)):
                        middle_high_data = {
                            column_mapping.get(middle_high_column_names[i], middle_high_column_names[i]): str(row[i].value) if row[i].value is not None else ''
                            for i in range(middle_high_start_index, middle_high_end_index + 1)
                        }
                        middle_high_data['import_date'] = datetime.now().strftime("%Y-%m-%d")  # Add the import_date value
                        user_data_list.append((
                            re.sub('\s.*', '', str(middle_high_data['assignment_date'])),
                            re.sub('\.0$|_\d{8}', '', str(middle_high_data['lead_number']))+'_%s'%datetime.now().strftime("%Y%m%d"),
                            middle_high_data['source'],
                            middle_high_data['url'],
                            middle_high_data['vat'],
                            re.sub('\.0$', '', middle_high_data['amount_phone_number'] or ''),
                            re.sub('\.0$', '', middle_high_data['amount_emails'] or ''),
                            file.filename,
                            middle_high_data['import_date'],
                            middle_high_data['import_date']+'_middle_high',
                            middle_high_data['local_spanish_seller'],
                            middle_high_data['cluster'],
                            middle_high_data['company_name'],
                            middle_high_data['product_count'],
                            middle_high_data['revenue'],
                            middle_high_data['telephone'],
                            middle_high_data['telephone1'],
                            middle_high_data['email'],
                            middle_high_data['email1'],
                            middle_high_data['lead_id'],
                        ))

        if 'Database' in wb.sheetnames:
            print('Database',len(user_data_list))
            ws = wb['Database']
            if ws.sheet_state == 'hidden':
                print('Database is hidden')
                pass
            else:
                column_names = [cell.value for cell in ws[1]]
                min_row = 2
                if 'Assignment Date' not in column_names:
                    column_names = [cell.value for cell in ws[2]]
                    min_row = 3

                if start_column not in column_names or end_column not in column_names:
                    return jsonify({'status': 'Invalid column names provided'})

                start_index = column_names.index(start_column)
                end_index = column_names.index(end_column)

                for row in ws.iter_rows(min_row=min_row):
                    # Check if any data is present in the row
                    if any(row[i].value for i in range(start_index, end_index + 1)):
                        user_data = {
                            column_mapping.get(column_names[i], column_names[i]): str(row[i].value) if row[i].value is not None else ''
                            for i in range(start_index, end_index + 1)
                        }
                        user_data['import_date'] = datetime.now().strftime("%Y-%m-%d")  # Add the import_date value
                        user_data_list.append((
                            re.sub('\s.*', '', str(user_data['assignment_date'])),
                            re.sub('\.0$|_\d{8}', '', str(user_data['lead_number']))+'_%s'%datetime.now().strftime("%Y%m%d"),
                            user_data['source'],
                            user_data['url'],
                            user_data['vat'],
                            re.sub('\.0$', '', user_data['amount_phone_number'] or ''),
                            re.sub('\.0$', '', user_data['amount_emails'] or ''),
                            file.filename,
                            user_data['import_date'],
                            '',#high_priority
                            user_data['local_spanish_seller'],
                            user_data['cluster'],
                            user_data['company_name'],
                            user_data['product_count'],
                            user_data['revenue'],
                            user_data['telephone'],
                            user_data['telephone1'],
                            user_data['email'],
                            user_data['email1'],
                            '',  # lead_id
                        ))
        #print(user_data_list)
        
        select_distinct_dates_query = """
            SELECT distinct trim(assignment_date) FROM miravia.miravia_table ORDER BY 1
        """
        cursor.execute(select_distinct_dates_query)
        distinct_dates = [row[0].strip() for row in cursor.fetchall()][:-8]

        select_exists_leads_query = """
            SELECT trim(regexp_replace(lead_number, '_(\d{8})$', '', 'g')), url FROM miravia.miravia_table
        """
        cursor.execute(select_exists_leads_query)
        exists_leads = [(row[0].strip(), row[1].strip()) for row in cursor.fetchall()]

        insert_query = """
            INSERT INTO miravia.miravia_table (
                assignment_date, lead_number, source, url, vat, amount_phone_number, amount_emails, file_name, import_date, high_priority,
                local_spanish_seller, cluster, company_name, product_count, revenue, telephone, telephone1, email, email1, lead_id
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        inserted_count = 0
        duplicate_count = 0
        batch_data = []
        for data in user_data_list:
            if data[0].strip() in distinct_dates or (re.sub('_\d{8}$', '', data[1]).strip(), data[3]) in exists_leads:
                duplicate_count += 1
            else:
                batch_data.append(data)
                inserted_count += 1

        cursor.executemany(insert_query, batch_data)
        conn.commit()

        select_table = """SELECT assignment_date,lead_number,source,lead_id,url,product_service,local_spanish_seller,
                            vat,cluster,l1,l2,company_name,revenue,product_count,product_desc_or_at_least_2_product_images,
                            amount_phone_number,telephone,telephone1,amount_emails,email,email1,feedback_final,qualification_feedback
                            FROM miravia.miravia_table where file_name= %s order by id asc limit 50""" 
        cursor.execute(select_table,(file.filename,))
        imported_data = [row for row in cursor.fetchall()]
        print('asdad',imported_data)
        
        return {
            'status': 'File imported successfully!',
            'rows_inserted': inserted_count,
            'duplicate_rows': duplicate_count,
            'total': len(user_data_list),
            'imported_data':imported_data,
        }
    except Exception as e:
        return jsonify({'status': 'Error occurred during file import', 'error': str(e)})

@app.route('/import-data')
def import_home():
    return render_template('import.html')

@app.route('/report-file-imported')
def report():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    cursor.execute("""SELECT file_name, TO_CHAR(import_date::date, 'DD/MM/YYYY') AS import_date, COUNT(*) AS total,
                    COUNT(CASE WHEN end_time <> '' THEN 1 END) AS done,
                    COUNT(CASE WHEN end_time = '' or end_time IS NULL THEN 1 END) AS pending,
                    COUNT(CASE WHEN is_exported <> '' THEN 1 END) AS exported
                FROM miravia.miravia_table
                GROUP BY file_name, import_date
                ORDER BY TO_DATE(import_date, 'YYYY-MM-DD') desc;""")
    import_rows = cursor.fetchall()

    conn.close()

    return render_template('file-imported-report.html', import_rows=import_rows)

@app.route('/report-user-records')
def report1():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    exclusion_conditions = {
        "product_service": ["N/A", "Only Services"],
        "local_spanish_seller": ["N"],
        "feedback_final": ["MAINTENANCE WEBSITE"],
        "qualification_feedback": ["INACTIVE PAGE - Mainteinance"],
        "url": ["", "\s|@|amazon\.|Lazada|miravia.es|etsy|ebay\.|amzn.to"],
    }

    exclusion_query = ""
    for column, values in exclusion_conditions.items():
        if column == "url":
            value_conditions = [f"url <> ''", f"url ~ '\w+\.\w+' ", f"url !~* '{'|'.join([x for x in values if x !=''])}'"]
            exclusion_query += f"AND ({' AND '.join(value_conditions)}) "
        else:
            value_conditions = [f"{column} <> '{value}'" for value in values]
            exclusion_query += f"AND ({' AND '.join(value_conditions)}) "

    month_duration = {
        1: (25, 21),
        2: (22, 21),
        3: (22, 21),
        4: (22, 21),
        5: (22, 21),
        6: (22, 21),
        7: (22, 21),
        8: (22, 21),
        9: (22, 21),
        10: (22, 23),
        11: (24, 23),
        12: (24, 22)
    }

    
    cursor.execute("SELECT email FROM miravia.miravia_users")
    rows = cursor.fetchall()
    emails = [row[0] for row in rows]

    select_paid_lines_query = f"""
        SELECT COUNT(*) FROM miravia.miravia_table
        WHERE end_time<>'' {exclusion_query}
    """
    
    cursor.execute(select_paid_lines_query)
    total_paid_lines = cursor.fetchone()[0]

    # Count total all lines
    cursor.execute(f"SELECT COUNT(*) FROM miravia.miravia_table WHERE end_time <> ''")
    total_all_lines = cursor.fetchone()[0]

    paid_lines_this_month = 0
    your_paid_lines_this_month = 0
    all_lines_this_month = 0
    your_all_lines_this_month = 0

    
    current_year = datetime.now().year
    current_time = datetime.now()

    for month, duration in month_duration.items():
        start_day, end_day = duration
        if month == 1:
            start_date = f"{current_year-1}-12-{start_day:02d}"
            end_date = f"{current_year}-01-{end_day:02d}"
        else:
            start_date = f"{current_year}-{month-1:02d}-{start_day:02d}"
            end_date = f"{current_year}-{month:02d}-{end_day:02d}"

        if datetime.strptime(start_date, "%Y-%m-%d") <= current_time <= datetime.strptime(end_date, "%Y-%m-%d"):
            selected_start_date = start_date
            selected_end_date = end_date
            break

    cursor.execute(f"SELECT pic, COUNT(*) FROM miravia.miravia_table WHERE TO_DATE(end_time, 'YYYY-mm-dd')::DATE >= DATE '{selected_start_date}'::DATE"
                    f" AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE <= DATE '{selected_end_date}'::DATE {exclusion_query} GROUP BY pic")
    rows = cursor.fetchall()
    your_paid_lines = {row[0]: row[1] for row in rows}
    your_paid_lines_json = json.dumps(your_paid_lines)


    paid_lines_this_month = sum(your_paid_lines.get(email, 0) for email in emails)

    
    cursor.execute(f"SELECT pic, COUNT(*) FROM miravia.miravia_table WHERE TO_DATE(end_time, 'YYYY-mm-dd')::DATE >= DATE '{selected_start_date}'::DATE"
                    f" AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE <= DATE '{selected_end_date}'::DATE GROUP BY pic")
    rows = cursor.fetchall()
    your_all_lines = {row[0]: row[1] for row in rows}
    your_all_lines_json  = json.dumps(your_all_lines )


    all_lines_this_month = sum(your_all_lines.get(email, 0) for email in emails)

    
    





    cursor.execute("SELECT regexp_replace(email, '\@.*', '', 'g') FROM miravia.miravia_users ORDER BY id ASC")
    rows = cursor.fetchall()
    users = users1 = [row[0] for row in rows]

    cursor.execute(f"SELECT regexp_replace(pic, '\@.*', '', 'g'), DATE_TRUNC('day', TO_TIMESTAMP(regexp_replace(end_time, '\s.*', '', 'g'), 'YYYY-MM-DD')), COUNT(*) "
                   f"FROM miravia.miravia_table "
                   f"WHERE end_time <> '' "
                   f"{exclusion_query} "
                   f"GROUP BY 1, 2 "
                   f"ORDER BY MIN(TO_TIMESTAMP(regexp_replace(end_time, '\s.*', '', 'g'), 'YYYY-MM-DD')) DESC")
    rows = cursor.fetchall()
    pic_in_day = [dict(pic=row[0], day=row[1].strftime('%d/%m/%Y'), count_in_day=row[2]) for row in rows]

    cursor.execute(f"SELECT regexp_replace(end_time, '\s.*', '', 'g'), COUNT(*) "
                   f"FROM miravia.miravia_table "
                   f"WHERE end_time <> '' "
                   f"{exclusion_query} "
                   f"GROUP BY 1 "
                   f"ORDER BY regexp_replace(end_time, '\s.*', '', 'g')::date DESC")
    rows = cursor.fetchall()
    day_totals = [dict(day=datetime.strftime(datetime.strptime(row[0], '%Y-%m-%d'), '%d/%m/%Y'), total_in_day=row[1]) for row in rows]

    cursor.execute(f"SELECT SUM(count) "
                   f"FROM (SELECT regexp_replace(pic, '\@.*', '', 'g'), COUNT(*) "
                   f"FROM miravia.miravia_table "
                   f"WHERE end_time <> '' "
                   f"{exclusion_query} "
                   f"GROUP BY 1, DATE_TRUNC('day', TO_TIMESTAMP(regexp_replace(end_time, '\s.*', '', 'g'), 'YYYY-MM-DD'))) as subquery")
    all_total_count = cursor.fetchone()[0]
    if all_total_count is None:
        all_total_count = 0

    cursor.execute("SELECT regexp_replace(pic, '\@.*', '', 'g'), DATE_TRUNC('day', TO_TIMESTAMP(regexp_replace(end_time, '\s.*', '', 'g'), 'YYYY-MM-DD')), COUNT(*), SUM(COUNT(*)) OVER () FROM miravia.miravia_table WHERE end_time <> '' GROUP BY 1, 2 ORDER BY MIN(TO_TIMESTAMP(regexp_replace(end_time, '\s.*', '', 'g'), 'YYYY-MM-DD')) DESC")
    rows = cursor.fetchall()
    pic_in_day1 = [dict(pic=row[0], day=row[1].strftime('%d/%m/%Y'), count_in_day=row[2]) for row in rows]

    cursor.execute("SELECT regexp_replace(end_time, '\s.*', '', 'g'), COUNT(*) FROM miravia.miravia_table WHERE end_time <> '' GROUP BY 1 ORDER BY regexp_replace(end_time, '\s.*', '', 'g')::date DESC")
    rows = cursor.fetchall()
    day_totals1 = [dict(day=datetime.strftime(datetime.strptime(row[0], '%Y-%m-%d'), '%d/%m/%Y'), total_in_day=row[1]) for row in rows]

    cursor.execute("SELECT SUM(count) FROM (SELECT regexp_replace(pic, '\@.*', '', 'g'), COUNT(*) FROM miravia.miravia_table WHERE end_time <> '' GROUP BY 1, DATE_TRUNC('day', TO_TIMESTAMP(regexp_replace(end_time, '\s.*', '', 'g'), 'YYYY-MM-DD'))) as subquery")
    all_total_count1 = cursor.fetchone()[0]
    if all_total_count1 is None:
        all_total_count1 = 0

    conn.close()

    return render_template('user-records-report.html', total_paid_lines=total_paid_lines,total_all_lines=total_all_lines,paid_lines_this_month=paid_lines_this_month,all_lines_this_month=all_lines_this_month,your_paid_lines_json=your_paid_lines_json,your_all_lines_json=your_all_lines_json,users=users, pic_in_day=pic_in_day, day_totals=day_totals, all_total_count=all_total_count,users1=users1, pic_in_day1=pic_in_day1, day_totals1=day_totals1, all_total_count1=all_total_count1)

@app.template_filter('get_count_in_day')
def get_count_in_day(user, day, pic_in_day):
    for item in pic_in_day:
        if item['pic'] == user and item['day'] == day:
            return item['count_in_day']
    return 0

@app.template_filter('get_total_count')
def get_total_count(day_total, pic_in_day):
    total_count = sum(item['count_in_day'] for item in pic_in_day if item['day'] == day_total['day'])
    return total_count

@app.template_filter('get_all_total_count')
def get_all_total_count(pic_in_day):
    all_total_count = sum(item['count_in_day'] for item in pic_in_day)
    return all_total_count

@app.template_filter('get_count_in_day1')
def get_count_in_day1(user, day, pic_in_day1):
    for item in pic_in_day1:
        if item['pic'] == user and item['day'] == day:
            return item['count_in_day']
    return 0


@app.template_filter('get_total_count1')
def get_total_count1(day_total, pic_in_day1):
    total_count = sum(item['count_in_day'] for item in pic_in_day1 if item['day'] == day_total['day'])
    return total_count

@app.template_filter('get_all_total_count1')
def get_all_total_count1(pic_in_day1):
    all_total_count1 = sum(item['count_in_day'] for item in pic_in_day1)
    return all_total_count1



@app.route('/')
def home():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    #thêm user thì thêm user_options
    username = request.form['username']
    password = request.form['password']
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    users_query = '''
                SELECT * FROM miravia.miravia_users
            '''
    cursor.execute(users_query,)
    rows = cursor.fetchall()
    users = [dict(email=row[0],password=row[1],role=row[2]) for row in rows]
    #print(users)
    
    # Perform login verification
    if username in [x['email'] for x in users]:
        if password in ''.join([x['password'] for x in users if x['email']==username]):
            return jsonify({'success': True,'user':username,'role':''.join([x['role'] for x in users if x['email']==username])})
        else:
            return jsonify({'success': False})
    else:
        return jsonify({'success': False})

@app.route('/form', methods = ['GET','POST'])
def index():

    categories_l1 = ["Audio","Bags and Travel","Beauty","Bedding & Bath","Cameras & Drones","Computers & Laptops","Data Storage","Electronics Accessories","Furniture & Organization","Groceries","Health","Household Supplies","Kid's Shoes and Clothing","Kitchen & Dining","Large Appliances","Laundry & Cleaning Equipment","Lighting & Décor","Men's Shoes and Clothing","Mobiles & Tablets","Mother & Baby","Outdoor & Garden","Pet Supplies","Small Appliances","Smart Devices","Sports & Outdoors","Sports Shoes and Clothing","Televisions & Videos","Tools & Home Improvement","Toys & Games","Watches Sunglasses Jewellery","Women's Shoes and Clothing"
]
    categories_l2 = ["Action Figures & Collectibles","Adult Diapers & Incontinence","Air Care","Air Conditioners","Arts & Crafts","Baby Gear","Baby Health Care","Baby Personal Care","Baby Safety","Baby & Toddler Toys","Bakery","Bakeware","Bath","Beauty Tools","Bedding","Beer, Wine & Spirits","Boy's Clothing","Boy's Shoes","Boys' Sports Accessories","Boys' Sports Bags","Boys' Sports Clothing","Boys' Sports Shoes","Breakfast Cereals & Spreads","Building Blocks","Build-in Oven","Chocolate, Snacks & Sweets","Cleaning Supplies","Cleaning Tools","Clothes Dryers","Clothing & Accessories","Coffee & Tea","Computer Accessories","Contact Lenses","Cookware","Cutlery","Dairy & Chilled","Desktops Computers","Diapering & Potty","Dinnerware","Dishwashers","Dishwashing","Disposables","Dolls & Accessories","Drinks","Drinkware","Drones","Electronic & Remote Control Toys","Exercise & Fitness Equipment","External hard Drives","External Solid State Drives","Eyeglasses","Feeding Essentials","Flash Drives","Food Staples & Cooking Essentials","Food Supplement","Fragrances","Freezers","Frozen","Fruit & Vegetables","Furniture","Gadgets","Gifts","Girl's Clothing","Girl's Shoes","Girls' Sports Accessories","Girls' Sports Bags","Girls' Sports Clothing","Girls' Sports Shoes","Hand Tools","Hardware","Headphones & Headsets","Hobbies & Entertainment","Home Audio","Home Décor","Household Sundries","Instant Camera","Internal Solid State Drives","Investment Precious Metal","Kids Bags","Kids Jewellery","Kids Watches","Kitchen Storage & Accessories","Kitchen Supplies","Kitchen & Table Linen","Kitchen Utensils","Laptops","Laundry Supplies","Laundry Tools & Accessories","Lawn & Garden","Learning & Education","Lighting","Live Animal","Makeup","Maternity Care","Meat & Seafood","Medical Supplies","Memory Cards","Men Bags","Men Clothing","Men Jewellery","Men's Care","Men Shoes","Men's Sports Accessories","Men's Sports Bags","Men's Sports Clothing","Men's Sports Shoes","Men Watches","Microwaves","Milk Formula & Baby Food","Network Components","Nursery","Outdoor","Outdoor Recreation","Pacifiers & Accessories","Paper & Tissue","Party Supplies","Personal Care","Personal Care Appliances","Pest Control","Pet Accessories","Pet Food","Pet Healthcare","Portable Speakers","Power Tools","Pretend Play, Costumes & Party","Refrigerators","Security Cameras & Systems","Serveware","Sexual Wellness","Skin Care","Small Household Appliances","Small Kitchen Appliances","Smartphones","Smart Televisions","Smartwatches & Accessories","Sports Toys & Outdoor Play","Storage & Organization","Sunglasses","Tablets","Team Merchandise/Fan Shop","Tool Storage & Shelving","Traditional Games","Travel","TV Accessories","Video","Video & Action Camcorder","Virtual Reality","Washing Machines","Wine Cellars","Women Bags","Women Clothing","Women Jewellery","Women Shoes","Women's Sports Accessories","Women's Sports Bags","Women's Sports Clothing","Women's Sports Shoes","Women Watches"]

    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    cursor.execute("SELECT count(*) FROM miravia.miravia_table where pic='' or pic is null")
    rows = cursor.fetchone()
    record_left = rows[0]
    
    pic = request.form.get('pic')
    #print(pic)

    cursor.execute("SELECT count(*) FROM miravia.miravia_table where pic= '%s' and end_time<>''" %pic)
    rows = cursor.fetchone()
    your_score = rows[0]

    # paid value------------------------------------------------------------
    exclusion_conditions = {
        "product_service": ["N/A", "Only Services"],
        "local_spanish_seller": ["N"],
        "feedback_final": ["MAINTENANCE WEBSITE"],
        "qualification_feedback": ["INACTIVE PAGE - Mainteinance"],
        "url": ["", "\s|@|amazon\.|Lazada|miravia.es|etsy|ebay\.|amzn.to"],
    }

    exclusion_query = ""
    for column, values in exclusion_conditions.items():
        if column == "url":
            value_conditions = [f"url <> ''", f"url ~ '\w+\.\w+' ", f"url !~* '{'|'.join([x for x in values if x !=''])}'"]
            exclusion_query += f"AND ({' AND '.join(value_conditions)}) "
        else:
            value_conditions = [f"{column} <> '{value}'" for value in values]
            exclusion_query += f"AND ({' AND '.join(value_conditions)}) "

    cursor.execute(f"SELECT COUNT(*) "
                f"FROM miravia.miravia_table "
                f"WHERE pic = '%s' "
                f"and end_time <> '' "
                f"{exclusion_query}"%pic)
    rows = cursor.fetchone()
    your_paid = rows[0]
    #print(your_score,your_paid)
    # ---------------------paid value------------------------------------------------
    

    conn.close()

    if request.method == 'POST':
        print('POST')
        return jsonify({'your_score':your_score,'your_paid':your_paid})

    return render_template('homepage.html', title='Miravia Home Input',categories_l1=categories_l1,categories_l2=categories_l2,record_left=record_left,your_score=your_score,your_paid=your_paid)

@app.route('/get_record_left')
def update_record_left():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    cursor.execute("SELECT count(*) FROM miravia.miravia_table where pic='' or pic is null")
    rows = cursor.fetchone()
    record_left = rows[0]

    conn.close()

    return {'record_left': record_left}

@app.route('/modal-api')
def modal_api():

    categories_l1 = ["Audio","Bags and Travel","Beauty","Bedding & Bath","Cameras & Drones","Computers & Laptops","Data Storage","Electronics Accessories","Furniture & Organization","Groceries","Health","Household Supplies","Kid's Shoes and Clothing","Kitchen & Dining","Large Appliances","Laundry & Cleaning Equipment","Lighting & Décor","Men's Shoes and Clothing","Mobiles & Tablets","Mother & Baby","Outdoor & Garden","Pet Supplies","Small Appliances","Smart Devices","Sports & Outdoors","Sports Shoes and Clothing","Televisions & Videos","Tools & Home Improvement","Toys & Games","Watches Sunglasses Jewellery","Women's Shoes and Clothing"
]
    categories_l2 = ["Action Figures & Collectibles","Adult Diapers & Incontinence","Air Care","Air Conditioners","Arts & Crafts","Baby Gear","Baby Health Care","Baby Personal Care","Baby Safety","Baby & Toddler Toys","Bakery","Bakeware","Bath","Beauty Tools","Bedding","Beer, Wine & Spirits","Boy's Clothing","Boy's Shoes","Boys' Sports Accessories","Boys' Sports Bags","Boys' Sports Clothing","Boys' Sports Shoes","Breakfast Cereals & Spreads","Building Blocks","Build-in Oven","Chocolate, Snacks & Sweets","Cleaning Supplies","Cleaning Tools","Clothes Dryers","Clothing & Accessories","Coffee & Tea","Computer Accessories","Contact Lenses","Cookware","Cutlery","Dairy & Chilled","Desktops Computers","Diapering & Potty","Dinnerware","Dishwashers","Dishwashing","Disposables","Dolls & Accessories","Drinks","Drinkware","Drones","Electronic & Remote Control Toys","Exercise & Fitness Equipment","External hard Drives","External Solid State Drives","Eyeglasses","Feeding Essentials","Flash Drives","Food Staples & Cooking Essentials","Food Supplement","Fragrances","Freezers","Frozen","Fruit & Vegetables","Furniture","Gadgets","Gifts","Girl's Clothing","Girl's Shoes","Girls' Sports Accessories","Girls' Sports Bags","Girls' Sports Clothing","Girls' Sports Shoes","Hand Tools","Hardware","Headphones & Headsets","Hobbies & Entertainment","Home Audio","Home Décor","Household Sundries","Instant Camera","Internal Solid State Drives","Investment Precious Metal","Kids Bags","Kids Jewellery","Kids Watches","Kitchen Storage & Accessories","Kitchen Supplies","Kitchen & Table Linen","Kitchen Utensils","Laptops","Laundry Supplies","Laundry Tools & Accessories","Lawn & Garden","Learning & Education","Lighting","Live Animal","Makeup","Maternity Care","Meat & Seafood","Medical Supplies","Memory Cards","Men Bags","Men Clothing","Men Jewellery","Men's Care","Men Shoes","Men's Sports Accessories","Men's Sports Bags","Men's Sports Clothing","Men's Sports Shoes","Men Watches","Microwaves","Milk Formula & Baby Food","Network Components","Nursery","Outdoor","Outdoor Recreation","Pacifiers & Accessories","Paper & Tissue","Party Supplies","Personal Care","Personal Care Appliances","Pest Control","Pet Accessories","Pet Food","Pet Healthcare","Portable Speakers","Power Tools","Pretend Play, Costumes & Party","Refrigerators","Security Cameras & Systems","Serveware","Sexual Wellness","Skin Care","Small Household Appliances","Small Kitchen Appliances","Smartphones","Smart Televisions","Smartwatches & Accessories","Sports Toys & Outdoor Play","Storage & Organization","Sunglasses","Tablets","Team Merchandise/Fan Shop","Tool Storage & Shelving","Traditional Games","Travel","TV Accessories","Video","Video & Action Camcorder","Virtual Reality","Washing Machines","Wine Cellars","Women Bags","Women Clothing","Women Jewellery","Women Shoes","Women's Sports Accessories","Women's Sports Bags","Women's Sports Clothing","Women's Sports Shoes","Women Watches"]

    return jsonify({'categories_l1':categories_l1,'categories_l2':categories_l2})

def fetch_users(sort_field=None):
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()
    sort_field = 'start_time'
    if sort_field is not None:
        cursor.execute(f"SELECT * FROM miravia.miravia_table where start_time<>'' and is_exported is null ORDER BY id desc")
    else:
        cursor.execute('SELECT * FROM miravia.miravia_table start_time<>'' and is_exported is null ORDER BY id desc')

    rows = cursor.fetchall()
    users = [dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],is_exported=row[29],file_name=row[30],note_for_qc=row[31]) for row in rows]
    
    cursor.close()
    conn.close()

    return users

@app.route('/database')
def table_data():
    users = fetch_users()
    users_json = json.dumps(users) 
    categories_l1 = ["Audio","Bags and Travel","Beauty","Bedding & Bath","Cameras & Drones","Computers & Laptops","Data Storage","Electronics Accessories","Furniture & Organization","Groceries","Health","Household Supplies","Kid's Shoes and Clothing","Kitchen & Dining","Large Appliances","Laundry & Cleaning Equipment","Lighting & Décor","Men's Shoes and Clothing","Mobiles & Tablets","Mother & Baby","Outdoor & Garden","Pet Supplies","Small Appliances","Smart Devices","Sports & Outdoors","Sports Shoes and Clothing","Televisions & Videos","Tools & Home Improvement","Toys & Games","Watches Sunglasses Jewellery","Women's Shoes and Clothing"
]
    categories_l2 = ["Action Figures & Collectibles","Adult Diapers & Incontinence","Air Care","Air Conditioners","Arts & Crafts","Baby Gear","Baby Health Care","Baby Personal Care","Baby Safety","Baby & Toddler Toys","Bakery","Bakeware","Bath","Beauty Tools","Bedding","Beer, Wine & Spirits","Boy's Clothing","Boy's Shoes","Boys' Sports Accessories","Boys' Sports Bags","Boys' Sports Clothing","Boys' Sports Shoes","Breakfast Cereals & Spreads","Building Blocks","Build-in Oven","Chocolate, Snacks & Sweets","Cleaning Supplies","Cleaning Tools","Clothes Dryers","Clothing & Accessories","Coffee & Tea","Computer Accessories","Contact Lenses","Cookware","Cutlery","Dairy & Chilled","Desktops Computers","Diapering & Potty","Dinnerware","Dishwashers","Dishwashing","Disposables","Dolls & Accessories","Drinks","Drinkware","Drones","Electronic & Remote Control Toys","Exercise & Fitness Equipment","External hard Drives","External Solid State Drives","Eyeglasses","Feeding Essentials","Flash Drives","Food Staples & Cooking Essentials","Food Supplement","Fragrances","Freezers","Frozen","Fruit & Vegetables","Furniture","Gadgets","Gifts","Girl's Clothing","Girl's Shoes","Girls' Sports Accessories","Girls' Sports Bags","Girls' Sports Clothing","Girls' Sports Shoes","Hand Tools","Hardware","Headphones & Headsets","Hobbies & Entertainment","Home Audio","Home Décor","Household Sundries","Instant Camera","Internal Solid State Drives","Investment Precious Metal","Kids Bags","Kids Jewellery","Kids Watches","Kitchen Storage & Accessories","Kitchen Supplies","Kitchen & Table Linen","Kitchen Utensils","Laptops","Laundry Supplies","Laundry Tools & Accessories","Lawn & Garden","Learning & Education","Lighting","Live Animal","Makeup","Maternity Care","Meat & Seafood","Medical Supplies","Memory Cards","Men Bags","Men Clothing","Men Jewellery","Men's Care","Men Shoes","Men's Sports Accessories","Men's Sports Bags","Men's Sports Clothing","Men's Sports Shoes","Men Watches","Microwaves","Milk Formula & Baby Food","Network Components","Nursery","Outdoor","Outdoor Recreation","Pacifiers & Accessories","Paper & Tissue","Party Supplies","Personal Care","Personal Care Appliances","Pest Control","Pet Accessories","Pet Food","Pet Healthcare","Portable Speakers","Power Tools","Pretend Play, Costumes & Party","Refrigerators","Security Cameras & Systems","Serveware","Sexual Wellness","Skin Care","Small Household Appliances","Small Kitchen Appliances","Smartphones","Smart Televisions","Smartwatches & Accessories","Sports Toys & Outdoor Play","Storage & Organization","Sunglasses","Tablets","Team Merchandise/Fan Shop","Tool Storage & Shelving","Traditional Games","Travel","TV Accessories","Video","Video & Action Camcorder","Virtual Reality","Washing Machines","Wine Cellars","Women Bags","Women Clothing","Women Jewellery","Women Shoes","Women's Sports Accessories","Women's Sports Bags","Women's Sports Clothing","Women's Sports Shoes","Women Watches"]
    
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    cursor.execute(f"SELECT email FROM miravia.miravia_users order by id asc")
    rows = cursor.fetchall()
    user_options = [row[0] for row in rows]

    return render_template('server_table.html', title='Miravia Database', users=users_json,categories_l1=categories_l1,categories_l2=categories_l2,user_options=user_options)#,columns=columns

@app.route('/api/modal')
def modal_show():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    # search filter
    lead_number = request.args.get('lead-number')
    url = re.sub('^WEB:','',request.args.get('url'))
    query = '''
            SELECT * FROM miravia.miravia_table
            WHERE lead_number = %s and url = %s
        '''
    cursor.execute(query, (lead_number,url,))
    rows = cursor.fetchall()
        
    users = [dict(
        id=row[0],
        assignment_date=row[1],
        pic=row[2],
        lead_number=row[3],
        source=row[4],
        url=url,
        product_service=row[6],
        is_local_spanish_seller=row[7],
        vat=row[8],
        cluster=row[9],
        l1=html.unescape(row[10] or ''),
        l2=html.unescape(row[11] or ''),
        company_name=html.unescape(row[12] or ''),
        revenue=row[13],
        product_count=row[14],
        product_desc_or_at_least_2_product_images=row[15],
        amount_phone_number=row[16],
        telephone=row[17],
        telephone1=row[18],
        amount_emails=row[19],
        email=row[20],
        email1=row[21],
        feedback_final=row[22],
        qualification_feedback=row[23],
        resolution_date=row[24],
        leads_id=row[25],
        prio=row[26],
        start_time=row[27],
        end_time=row[28],
        is_exported=row[29],
        file_name=row[30],
        note_for_qc=row[31]
    ) for row in rows]
    cursor.close()
    conn.close()

    response = users[0]
    return jsonify(response)

@app.route('/api/data')
def table_api_data():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    # search filter
    search = request.args.get('search[value]')
    if search:
        search_query = f"%{search}%" if search else None
        query = '''
            SELECT * FROM miravia.miravia_table
            WHERE pic ~ %s and is_exported is null or lead_number ~ %s and is_exported is null or assignment_date ~ %s and is_exported is null or feedback_final ~ %s and is_exported is null or qualification_feedback ~ %s and is_exported is null
        '''
        cursor.execute(query, (search_query,search_query,search_query,))
        rows = cursor.fetchall()
        users = [dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],is_exported=row[29],file_name=row[30],note_for_qc=row[31]) for row in rows]

    # Get the latest date from the "end_time" column
    # latest_date_query = '''
    #     SELECT start_time FROM miravia.miravia_table WHERE start_time <> '' ORDER BY id DESC LIMIT 1
    # '''
    # cursor.execute(latest_date_query)
    
    # # load all latest date
    # latest_date = re.sub('\s.*', '', cursor.fetchone()[0])

    # date filter
    date_picker = request.args.get('datePicker')
    starttime_picker = request.args.get('starttimePicker')
    endtime_picker = request.args.get('endtimePicker')
    datashow_filter = request.args.get('datashow_filter')

    # Update the date filter to use the latest date
    all_query = '''
        SELECT * FROM miravia.miravia_table
        WHERE start_time<>'' and is_exported is null order by id desc
    '''
    cursor.execute(all_query)
    rows = cursor.fetchall()
    users = [dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],is_exported=row[29],file_name=row[30],note_for_qc=row[31]) for row in rows]
    
    if datashow_filter:
        if datashow_filter == 'done':
            data_show_query = '''
                SELECT * FROM miravia.miravia_table
                WHERE is_exported is null and start_time <>'' order by id desc
            '''
            cursor.execute(data_show_query, (datashow_filter,))
        elif datashow_filter == 'pending':
            data_show_query = '''
                SELECT * FROM miravia.miravia_table
                WHERE is_exported is null and (pic = '' or pic is null) order by id desc
            '''
            cursor.execute(data_show_query, (datashow_filter,))

        elif datashow_filter == 'exported':
            data_show_query = '''
                SELECT * FROM miravia.miravia_table
                WHERE is_exported ~'\w' order by id desc
            '''
            cursor.execute(data_show_query, (datashow_filter,))

        elif datashow_filter == 'total':
            data_show_query = '''
                SELECT * FROM miravia.miravia_table
                 order by id desc
            '''
            cursor.execute(data_show_query, (datashow_filter,))

        rows = cursor.fetchall()
        users = [dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],is_exported=row[29],file_name=row[30],note_for_qc=row[31]) for row in rows]
    
    if starttime_picker:
        starttime_query = '''
            SELECT * FROM miravia.miravia_table
            WHERE is_exported is null and start_time <>'' and regexp_replace(start_time,'.*\s|\:.*','','g')::integer >= %s
        '''
        cursor.execute(starttime_query, (starttime_picker,))
        rows = cursor.fetchall()
        users = [dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],is_exported=row[29],file_name=row[30],note_for_qc=row[31]) for row in rows]
        if endtime_picker:
            endtime_query = '''
                SELECT * FROM miravia.miravia_table
                WHERE is_exported is null and start_time <>'' and regexp_replace(start_time,'.*\s|\:.*','','g')::integer >= %s and end_time<>'' and regexp_replace(end_time,'.*\s|\:.*','','g')::integer <= %s
            '''
            cursor.execute(endtime_query, (starttime_picker,endtime_picker,))
            rows = cursor.fetchall()
            users = [dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],is_exported=row[29],file_name=row[30],note_for_qc=row[31]) for row in rows]
    else:
        if endtime_picker:
            endtime_query = '''
                SELECT * FROM miravia.miravia_table
                WHERE is_exported is null and start_time <>''  and regexp_replace(end_time,'.*\s|\:.*','','g')::integer <= %s
            '''
            cursor.execute(endtime_query, (endtime_picker,))
            rows = cursor.fetchall()
            users = [dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],is_exported=row[29],file_name=row[30],note_for_qc=row[31]) for row in rows]
   
    
    if date_picker:
        date_query = '''
            SELECT * FROM miravia.miravia_table
            WHERE is_exported is null and start_time ~ %s
        '''
        cursor.execute(date_query, (date_picker,))
        rows = cursor.fetchall()
        users = [dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],is_exported=row[29],file_name=row[30],note_for_qc=row[31]) for row in rows]

        if starttime_picker:
            starttime_query = '''
                SELECT * FROM miravia.miravia_table
                WHERE is_exported is null and start_time ~ %s and regexp_replace(start_time,'.*\s|\:.*','','g')::integer >= %s
            '''
            cursor.execute(starttime_query, (date_picker,starttime_picker,))
            rows = cursor.fetchall()
            users = [dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],is_exported=row[29],file_name=row[30],note_for_qc=row[31]) for row in rows]
            if endtime_picker:
                endtime_query = '''
                    SELECT * FROM miravia.miravia_table
                    WHERE is_exported is null and start_time ~ %s and regexp_replace(start_time,'.*\s|\:.*','','g')::integer >= %s and end_time<>'' and regexp_replace(end_time,'.*\s|\:.*','','g')::integer <= %s
                '''
                cursor.execute(endtime_query, (date_picker,starttime_picker,endtime_picker,))
                rows = cursor.fetchall()
                users = [dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],is_exported=row[29],file_name=row[30],note_for_qc=row[31]) for row in rows]
        else:
            if endtime_picker:
                endtime_query = '''
                    SELECT * FROM miravia.miravia_table
                    WHERE is_exported is null and start_time ~ %s  and regexp_replace(end_time,'.*\s|\:.*','','g')::integer <= %s
                '''
                cursor.execute(endtime_query, (date_picker,endtime_picker,))
                rows = cursor.fetchall()
                users = [dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],is_exported=row[29],file_name=row[30],note_for_qc=row[31]) for row in rows]
   
    # count the filtered records
    total_filtered = len(users)

    # sorting
    order = []
    i = 0
    while True:
        col_index = request.args.get(f'order[{i}][column]')
        if col_index is None:
            break
        col_name = request.args.get(f'columns[{col_index}][data]')
        if col_name not in ['start_time','lead_number', 'source', 'url', 'product_service', 'local_spanish_seller', 'vat', 'cluster', 'l1', 'l2', 'company_name', 'revenue', 'product_count', 'product_desc_or_at_least_2_product_images', 'amount_phone_number', 'telephone', 'telephone1', 'amount_emails', 'email', 'email1', 'feedback_final', 'qualification_feedback', 'pic']:
            col_name = 'pic'
        descending = request.args.get(f'order[{i}][dir]') == 'desc'
        col = col_name
        if descending:
            col = f'{col} DESC'
        order.append(col)
        i += 1
    if order:
        order_str = ', '.join(order)
        query += f' ORDER BY {order_str}'
        if date_picker:
            date_query += f' ORDER BY {order_str}'
        #if latest_date:
        #    latest_date_query += f' ORDER BY {order_str}'

    # pagination
    start = request.args.get('start', type=int)
    length = request.args.get('length', type=int)
    if start is not None and length is not None:
        offset = start
        limit = length
        if date_picker:
            date_query += ' OFFSET %s LIMIT %s'
            cursor.execute(date_query, (date_picker, offset, limit))
        else:
            query += ' OFFSET %s LIMIT %s'
            cursor.execute(query, (search_query, offset, limit))
        rows = cursor.fetchall()
        users = [dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],is_exported=row[29],file_name=row[30],note_for_qc=row[31]) for row in rows]

    cursor.close()
    conn.close()

    # response
    return {
        'data': users,
        'recordsFiltered': total_filtered,
        'recordsTotal': len(users)
    }

@app.route('/', methods=['POST'])
def input_form():
    if request.method == 'POST':
        # Get the form data
        end_time = request.form.get('end-time')
        pic = request.form.get('pic')
        lead_number = request.form.get('lead-number')
        url = request.form.get('url')
        product_service = request.form.get('product-service')
        local_spanish_seller = request.form.get('local_spanish_seller')
        vat = request.form.get('vat')
        cluster = request.form.get('cluster')
        l1 = request.form.get('l1')
        l2 = request.form.get('l2')
        company_name = request.form.get('company-name')
        revenue = request.form.get('revenue')
        product_count = request.form.get('product-count')
        product_desc_or_at_least_2_product_images = request.form.get('product_desc_or_at_least_2_product_images')
        amount_phone_number = request.form.get('amount-of-phonenumbers')
        telephone = request.form.get('telephone')
        telephone1 = request.form.get('telephone1')
        amount_emails = request.form.get('amount-of-emails')
        email = request.form.get('email')
        email1 = request.form.get('email1')
        feedback_final = request.form.get('feedback-final')
        qualification_feedback = request.form.get('qualification-feedback')
        note_for_qc = request.form.get('note')
        
        # Connect to the database
        conn = psycopg2.connect(app.config['DATABASE_URI'])
        cursor = conn.cursor()
        
        if pic == '' or lead_number == '':
            return jsonify({'error': 'Missing Lead Number or PIC'}), 500
        else:
            select_query = """
            SELECT url, pic
            FROM miravia.miravia_table
            WHERE lead_number = %s
            """
            cursor.execute(select_query, (lead_number,))
            existing_data = cursor.fetchone()

            if existing_data:
                existing_url, existing_pic = existing_data
                if pic != existing_pic and existing_pic!= '':
                    # URL and PIC mismatch error
                    return jsonify({'error': 'URL and PIC mismatch','existing_url':existing_url,'existing_pic':existing_pic,'url':url}), 500
                else:
                    datetime_string = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    # Update existing data
                    update_query = """
                        UPDATE miravia.miravia_table
                        SET url = %s, product_service = %s, local_spanish_seller = %s, vat = %s,
                            cluster = %s, l1 = %s, l2 = %s, company_name = %s, revenue = %s,
                            product_count = %s, product_desc_or_at_least_2_product_images = %s,
                            amount_phone_number = %s, telephone = %s, telephone1 = %s,
                            amount_emails = %s, email = %s, email1 = %s,
                            feedback_final = %s, qualification_feedback = %s,
                            end_time = %s, note_for_qc = %s
                        WHERE pic = %s AND lead_number = %s
                        """
                    cursor.execute(update_query, (
                        url or '', product_service or '', local_spanish_seller or '',
                        vat or '', cluster or '', html.escape(l1 or ''), html.escape(l2 or ''), company_name or '', revenue or '',
                        product_count or '', product_desc_or_at_least_2_product_images or '', amount_phone_number or '',
                        telephone or '', telephone1 or '', amount_emails or '', email or '', email1 or '',
                        feedback_final or '', qualification_feedback or '',
                        end_time !='' and end_time is not None and end_time or datetime_string, note_for_qc,
                        pic, lead_number
                    ))

            else:
                # Create an SQL INSERT query
                insert_query = """
                INSERT INTO miravia.miravia_table (
                    pic, lead_number, url, product_service, local_spanish_seller, vat, cluster, l1, l2,
                    company_name, revenue, product_count, product_desc_or_at_least_2_product_images,
                    amount_phone_number, telephone, telephone1, amount_emails, email, email1,
                    feedback_final, qualification_feedback, note_for_qc
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """

                # Execute the INSERT query with the form data
                cursor.execute(insert_query, (
                    pic or '', lead_number or '', url or '', product_service or '', local_spanish_seller or '',
                    vat or '', cluster or '', html.escape(l1 or ''), html.escape(l2 or ''), company_name or '', revenue or '',
                    product_count or '', product_desc_or_at_least_2_product_images or '', amount_phone_number or '',
                    telephone or '', telephone1 or '', amount_emails or '', email or '', email1 or '',
                    feedback_final or '', qualification_feedback or '', note_for_qc or ''
                ))

            # Commit the changes to the database
            conn.commit()

            # Close the cursor and the database connection
            cursor.close()
            conn.close()

        # Render the initial form template
        return render_template('homepage.html')
    
@app.route('/delete-pic', methods=['GET', 'POST'])
def delete_pic():
    if request.method == 'POST':
        # Get the form data
        end_time = request.form.get('end-time')
        pic = request.form.get('pic')
        lead_number = request.form.get('lead-number')
        url = request.form.get('url')
        product_service = request.form.get('product-service')
        local_spanish_seller = request.form.get('local_spanish_seller')
        vat = request.form.get('vat')
        cluster = request.form.get('cluster')
        l1 = request.form.get('l1')
        l2 = request.form.get('l2')
        company_name = request.form.get('company-name')
        revenue = request.form.get('revenue')
        product_count = request.form.get('product-count')
        product_desc_or_at_least_2_product_images = request.form.get('product_desc_or_at_least_2_product_images')
        amount_phone_number = request.form.get('amount-of-phonenumbers')
        telephone = request.form.get('telephone')
        telephone1 = request.form.get('telephone1')
        amount_emails = request.form.get('amount-of-emails')
        email = request.form.get('email')
        email1 = request.form.get('email1')
        feedback_final = request.form.get('feedback-final')
        qualification_feedback = request.form.get('qualification-feedback')
        
        # Connect to the database
        conn = psycopg2.connect(app.config['DATABASE_URI'])
        cursor = conn.cursor()
        
        if pic == '' or url == '':
            return jsonify({'error': 'Missing URL or PIC'}), 500
        else:
            #print(pic,url,lead_number)
            # Update existing data
            update_query = """
                UPDATE miravia.miravia_table
                SET pic='' , start_time = '', end_time=''
                WHERE pic = %s AND url = %s AND lead_number = %s
                """
            cursor.execute(update_query, (
                pic, url, lead_number,
            ))

            conn.commit()
            # Close the cursor and the database connection
            cursor.close()
            conn.close()

        # Render the initial form template
        return render_template('homepage.html')



@app.route('/get-domain-data', methods=['GET'])
def get_domain_data():
    # Connect to the PostgreSQL database using psycopg2
    # Execute a SELECT query to retrieve the data you need
    # Fetch the data from the cursor

    # Example code to fetch data from the database
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    cursor.execute("SELECT count(*) FROM miravia.miravia_table where pic='' or pic is null")
    rows = cursor.fetchone()
    record_left = rows[0]

    pic = request.args.get('pic')
    leadnumbergo = request.args.get('leadnumbergo')

    cursor.execute("SELECT count(*) FROM miravia.miravia_table where pic= '%s' and end_time<>''" %pic)
    rows = cursor.fetchone()
    your_score = rows[0]

    # paid value------------------------------------------------------------
    exclusion_conditions = {
        "product_service": ["N/A", "Only Services"],
        "local_spanish_seller": ["N"],
        "feedback_final": ["MAINTENANCE WEBSITE"],
        "qualification_feedback": ["INACTIVE PAGE - Mainteinance"],
        "url": ["", "\s|@|amazon\.|Lazada|miravia.es|etsy|ebay\.|amzn.to"],
    }

    exclusion_query = ""
    for column, values in exclusion_conditions.items():
        if column == "url":
            value_conditions = [f"url <> ''", f"url ~ '\w+\.\w+' ", f"url !~* '{'|'.join([x for x in values if x !=''])}'"]
            exclusion_query += f"AND ({' AND '.join(value_conditions)}) "
        else:
            value_conditions = [f"{column} <> '{value}'" for value in values]
            exclusion_query += f"AND ({' AND '.join(value_conditions)}) "

    cursor.execute(f"SELECT COUNT(*) "
                f"FROM miravia.miravia_table "
                f"WHERE pic = %s "
                f"AND end_time <> '' "
                f"{exclusion_query}",
                (pic,))
    rows = cursor.fetchone()
    your_paid = rows[0]

    

    # select exists lead number not done of pic
    select_query = """
    SELECT *
    FROM miravia.miravia_table
    WHERE lead_number <> '' AND pic = %s AND (end_time = '' OR end_time IS NULL)
    ORDER BY id ASC
    LIMIT 1
    """
    cursor.execute(select_query, (pic,))
    row = cursor.fetchone()

    if not row:
        select_query = """
            SELECT *
            FROM miravia.miravia_table
            WHERE pic = '' OR pic IS NULL
            ORDER BY 
                CASE 
                    WHEN high_priority ~ 'high_prio' THEN 1
                    WHEN high_priority ~ 'middle_high' THEN 2
                    ELSE 3
                END,
                id ASC
            LIMIT 1
        """
        cursor.execute(select_query)
        row = cursor.fetchone()
    if leadnumbergo:
        select_query = """
            SELECT url, pic
            FROM miravia.miravia_table
            WHERE lead_number = %s
            """
        cursor.execute(select_query, (leadnumbergo,))
        existing_data = cursor.fetchone()

        if existing_data:
            existing_url, existing_pic = existing_data
            if existing_pic!= '' and existing_pic is not None and pic != existing_pic:
                # URL and PIC mismatch error
                return jsonify({'error': 'Lead Number and PIC already existed!','existing_url':existing_url,'existing_pic':existing_pic,'lead_number':leadnumbergo}), 500
            else:
                cursor.execute("""
                    SELECT *
                    FROM miravia.miravia_table
                    WHERE lead_number = '%s'
                """ % (leadnumbergo,))
                row = cursor.fetchone()
    
    if row:
        datetime_string = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Update the 'start_time' field with the current datetime
        update_query = "UPDATE miravia.miravia_table SET start_time = %s WHERE url = %s and lead_number = %s"
        cursor.execute(update_query, (datetime_string, row[5], row[3],))

        conn.commit()  # Commit the transaction
        conn.close()
        data = dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28],record_left=record_left,your_score=your_score,your_paid=your_paid)

        return jsonify(data)
    else:
        return jsonify({'error': 'No data found in the database'}), 500
    
@app.route('/get-domain-data-save', methods=['GET'])
def get_domain_data_save():
    # Connect to the PostgreSQL database using psycopg2
    # Execute a SELECT query to retrieve the data you need
    # Fetch the data from the cursor

    # Example code to fetch data from the database
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()
    
    
    pic = request.args.get('pic')
    
    select_query = """
            SELECT *
            FROM miravia.miravia_table
            WHERE lead_number <>'' and pic = %s and end_time='' or lead_number <>'' and pic = %s and end_time is null order by id asc
            """
    cursor.execute(select_query, (pic,pic))
    row = cursor.fetchone()
    if not row:
        cursor.execute("""SELECT *
                       FROM miravia.miravia_table WHERE pic='' or pic is null ORDER BY id asc LIMIT 1""")
        row = cursor.fetchone()
    if row:
        datetime_string = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Update the 'start_time' field with the current datetime
        update_query = "UPDATE miravia.miravia_table SET start_time = %s WHERE url = %s AND lead_number = %s"
        cursor.execute(update_query, (datetime_string, row[5], row[3],))

        conn.commit()  # Commit the transaction

        data = dict(id=row[0],assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28])

        return jsonify(data)
    else:
        return jsonify({'error': 'No data found in the database'}), 500
    
@app.route('/get-domain-data-save-only', methods=['GET'])
def get_domain_data_save_only():
    # Connect to the PostgreSQL database using psycopg2
    # Execute a SELECT query to retrieve the data you need
    # Fetch the data from the cursor

    # Example code to fetch data from the database
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()
    
    pic = request.args.get('pic')
    
    select_query = """
            SELECT *
            FROM miravia.miravia_table
            WHERE lead_number <>'' and pic = %s and end_time='' or lead_number <>'' and pic = %s and end_time is null order by id asc
            """
    cursor.execute(select_query, (pic,pic))
    row = cursor.fetchone()
    if not row:
        cursor.execute("""SELECT *
                       FROM miravia.miravia_table WHERE pic='' or pic is null ORDER BY id asc LIMIT 1""")
        row = cursor.fetchone()
    if row:
        datetime_string = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Update the 'start_time' field with the current datetime
        update_query = "UPDATE miravia.miravia_table SET start_time = %s WHERE url = %s AND lead_number = %s"
        cursor.execute(update_query, (datetime_string, row[5], row[3],))

        conn.commit()  # Commit the transaction

        data = dict(assignment_date=row[1], pic=row[2], lead_number=row[3], source=row[4], url=row[5], product_service=row[6], is_local_spanish_seller=row[7], vat=row[8], cluster=row[9], l1=row[10], l2=row[11], company_name=row[12], revenue=row[13], product_count=row[14], product_desc_or_at_least_2_product_images=row[15], amount_phone_number=row[16], telephone=row[17], telephone1=row[18], amount_emails=row[19], email=row[20], email1=row[21], feedback_final=row[22], qualification_feedback=row[23], resolution_date=row[24], leads_id=row[25], prio=row[26], start_time=row[27], end_time=row[28])

        return jsonify(data)
    else:
        return jsonify({'error': 'No data found in the database'}), 500


@app.route('/update-pic', methods=['POST'])
def update_pic():
    # Connect to the PostgreSQL database using psycopg2
    # Execute an UPDATE query to update the 'pic' column based on the 'url' column

    # Retrieve the URL and PIC values from the request
    lead_number = request.json.get('lead_number')
    pic = request.json.get('pic')

    # Example code to update the 'pic' column in the database based on the 'url' value
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    cursor.execute("UPDATE miravia.miravia_table SET pic = %s WHERE lead_number = %s", (pic, lead_number))
    conn.commit()

    return jsonify({'message': 'PIC updated successfully','PIC':pic,'Lead Number':lead_number})

@app.route('/select-l1', methods=['GET'])
def select_l1():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()
    cluster = request.args.get('cluster')
    cursor.execute("select distinct l1 from miravia.miravia_cluster where cluster='%s' order by 1 asc"%cluster)

    rows = cursor.fetchall()
    l1_list = []
    if rows:
        for row in rows:
            l1_list.append(row[0])

        return jsonify(l1_list)
    else:
        return jsonify({'error': 'No L1 data found!'}), 500

@app.route('/select-l2', methods=['GET'])
def select_l2():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()
    l1 = request.args.get('l1')
    cursor.execute("select distinct l2 from miravia.miravia_cluster where l1=%s order by 1 asc", (l1,))

    rows = cursor.fetchall()
    l2_list = []
    if rows:
        for row in rows:
            l2_list.append(row[0])

        return jsonify(l2_list)
    else:
        return jsonify({'error': 'No L2 data found!'}), 500


@app.route('/export_report_file')
def export_report_file():
    file_name = request.args.get('file_name', '')

    # Connect to the database
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    if 'High Priority' in file_name:
        # Query the database for the rows with the specified file_name
        cursor.execute(f"""SELECT pic,file_name,note_for_qc,assignment_date,lead_number,source,lead_id,url,product_service,local_spanish_seller,
                            vat,cluster,l1,l2,company_name,revenue,product_count,product_desc_or_at_least_2_product_images,
                            amount_phone_number,telephone,telephone1,amount_emails,email,email1,feedback_final,qualification_feedback
                            FROM miravia.miravia_table where file_name = '{file_name}' order by id asc""")
        wb = openpyxl.load_workbook('[High Priority]_Template.xlsx')
    else:
        # Query the database for the rows with the specified file_name
        cursor.execute(f"""SELECT pic,file_name,note_for_qc,assignment_date,lead_number,source,url,product_service,local_spanish_seller,
                            vat,cluster,l1,l2,company_name,revenue,product_count,product_desc_or_at_least_2_product_images,
                            amount_phone_number,telephone,telephone1,amount_emails,email,email1,feedback_final,qualification_feedback
                            FROM miravia.miravia_table where file_name = '{file_name}' order by id asc""")
        wb = openpyxl.load_workbook('Miravia_Template.xlsx')
    data = cursor.fetchall()
    col_names = [desc[0] for desc in cursor.description]

    # Get the current day
    current_day = datetime.strftime(datetime.strptime(datetime.now().strftime("%Y/%m/%d"), '%Y/%m/%d'), '%d/%m/%Y')

    # Insert "Exported Day" column at the first position in col_names
    col_names.insert(0, 'Exported Day')

    # Insert "current_day" value at the first position in each row of data
    modified_data = [[current_day] + list(row) for row in data]

    # Load the Excel template
    
    ws = wb.active

    # Write data to the Excel template
    rowid = 2
    for r, row in enumerate(modified_data):
        for c, col in enumerate(row):
            try:
                if c == col_names.index('assignment_date'):
                    col = col is not None and datetime.strftime(datetime.strptime(col, '%Y-%m-%d'),'%d/%m/%Y') or ''
                if c == col_names.index('amount_phone_number'):
                    col = col is not None and col != '' and int(col) or 0
                if c == col_names.index('amount_emails'):
                    col = col is not None and col != '' and int(col) or 0
                if c == col_names.index('l1') or c == col_names.index('l2') or c == col_names.index('company_name'):
                    col = col is not None and col != '' and html.unescape(col) or col
                _ = ws.cell(column=c+1, row=r+rowid, value=col)
            except Exception as e:
                print(e)
                pass

    # Save the modified Excel file to a BytesIO object
    file_data = io.BytesIO()
    wb.save(file_data)
    file_data.seek(0)
    file_name = file_name.replace('.xlsx','')
    
    response = make_response(file_data.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
    return response


@app.route('/export_to_xlsx')
def export_to_xlsx():
    day = datetime.now().strftime("%Y%m%d")

    limit = request.args.get('limit','')
    template = request.args.get('template','')
    # Connect to the database
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    if template == 'database':
        if limit:
            cursor.execute(f"""SELECT pic,file_name,note_for_qc,assignment_date,lead_number,source,url,product_service,local_spanish_seller,
                            vat,cluster,l1,l2,company_name,revenue,product_count,product_desc_or_at_least_2_product_images,
                            amount_phone_number,telephone,telephone1,amount_emails,email,email1,feedback_final,qualification_feedback
                            FROM miravia.miravia_table where is_exported is null and end_time<>'' and (high_priority='' or high_priority is null) ORDER BY id ASC LIMIT {limit}""")
        else:
            cursor.execute("""SELECT pic,file_name,note_for_qc,assignment_date,lead_number,source,url,product_service,local_spanish_seller,
                            vat,cluster,l1,l2,company_name,revenue,product_count,product_desc_or_at_least_2_product_images,
                            amount_phone_number,telephone,telephone1,amount_emails,email,email1,feedback_final,qualification_feedback
                            FROM miravia.miravia_table where is_exported is null and end_time<>'' and (high_priority='' or high_priority is null) ORDER BY id ASC""")
    if template == 'highprio':
        if limit:
            cursor.execute(f"""SELECT pic,file_name,note_for_qc,assignment_date,lead_number,source,lead_id,url,product_service,local_spanish_seller,
                            vat,cluster,l1,l2,company_name,revenue,product_count,product_desc_or_at_least_2_product_images,
                            amount_phone_number,telephone,telephone1,amount_emails,email,email1,feedback_final,qualification_feedback
                            FROM miravia.miravia_table where is_exported is null and end_time<>'' and high_priority~'high_prio' ORDER BY id ASC LIMIT {limit}""")
        else:
            cursor.execute("""SELECT pic,file_name,note_for_qc,assignment_date,lead_number,source,lead_id,url,product_service,local_spanish_seller,
                            vat,cluster,l1,l2,company_name,revenue,product_count,product_desc_or_at_least_2_product_images,
                            amount_phone_number,telephone,telephone1,amount_emails,email,email1,feedback_final,qualification_feedback
                            FROM miravia.miravia_table where is_exported is null and end_time<>'' and high_priority~'high_prio' ORDER BY id ASC""")
    
    if template == 'middlehigh':
        if limit:
            cursor.execute(f"""SELECT pic,file_name,note_for_qc,assignment_date,lead_number,source,lead_id,url,product_service,local_spanish_seller,
                            vat,cluster,l1,l2,company_name,revenue,product_count,product_desc_or_at_least_2_product_images,
                            amount_phone_number,telephone,telephone1,amount_emails,email,email1,feedback_final,qualification_feedback
                            FROM miravia.miravia_table where is_exported is null and end_time<>'' and high_priority~'middle_high' ORDER BY id ASC LIMIT {limit}""")
        else:
            cursor.execute("""SELECT pic,file_name,note_for_qc,assignment_date,lead_number,source,lead_id,url,product_service,local_spanish_seller,
                            vat,cluster,l1,l2,company_name,revenue,product_count,product_desc_or_at_least_2_product_images,
                            amount_phone_number,telephone,telephone1,amount_emails,email,email1,feedback_final,qualification_feedback
                            FROM miravia.miravia_table where is_exported is null and end_time<>'' and high_priority~'middle_high' ORDER BY id ASC""")
        
    data = cursor.fetchall()
    col_names = [desc[0] for desc in cursor.description]

    # Get the current day
    current_day = datetime.strftime(datetime.strptime(datetime.now().strftime("%Y/%m/%d"), '%Y/%m/%d'), '%d/%m/%Y')

    # Insert "Exported Day" column at the first position in col_names
    col_names.insert(0, 'Exported Day')

    # Insert "current_day" value at the first position in each row of data
    modified_data = [[current_day] + list(row) for row in data]

    # Load the Excel template
    if template == 'database':
        wb = openpyxl.load_workbook('Miravia_Template.xlsx')
        ws = wb.active
    if template == 'middlehigh':
        wb = openpyxl.load_workbook('[Middle High]_Template.xlsx')
        ws = wb.active
    if template == 'highprio':
        wb = openpyxl.load_workbook('[High Priority]_Template.xlsx')
        ws = wb.active

    # Write data to the Excel template
    rowid = 2
    for r, row in enumerate(modified_data):
        for c, col in enumerate(row):
            try:
                if c == col_names.index('assignment_date'):
                    col = col is not None and datetime.strftime(datetime.strptime(col, '%Y-%m-%d'),'%d/%m/%Y') or ''
                if c == col_names.index('amount_phone_number'):
                    col = col is not None and col != '' and int(col) or 0
                if c == col_names.index('amount_emails'):
                    col = col is not None and col != '' and int(col) or 0
                if c == col_names.index('l1') or c == col_names.index('l2') or c == col_names.index('company_name'):
                    col = col is not None and col != '' and html.unescape(col) or col
                _ = ws.cell(column=c+1, row=r+rowid, value=col)
            except Exception as e:
                print(e)
                pass

    # Render the lead_number list
    lead_number_list = [str(row[col_names.index('lead_number') - 1]) for row in data]
    #print('export_to_xlsx', lead_number_list)
    # Save the modified Excel file to a BytesIO object
    file_data = BytesIO()
    wb.save(file_data)
    file_data.seek(0)

    # Send the file to the client
    response = make_response(file_data.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    if template == 'database':
        response.headers['Content-Disposition'] = ('attachment; filename=Miravia_[Database]_%s.xlsx' % day)
    if template == 'prio':
        response.headers['Content-Disposition'] = ('attachment; filename=Miravia_[High Priority]_%s.xlsx' % day)

    # Add the lead_number list to the response as a JSON
    response.headers['Lead-Number-List'] = json.dumps(lead_number_list)

    return response

@app.route('/get_lead_number_list', methods=['GET'])
def get_lead_number_list():
    global lead_number_list

    return jsonify({'lead_number_list': lead_number_list})

@app.route('/update_is_exported', methods=['POST'])
def update_is_exported():
    #global lead_number_list
    day = datetime.now().strftime("%Y%m%d")
    lead_number_list = request.json
    #print('update_is_exported:', lead_number_list)
    # Connect to the database
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    # Update the is_exported column for the lead numbers in lead_number_list
    for k,lead_numbers in lead_number_list.items():
        for lead_number in lead_numbers:
            cursor.execute("UPDATE miravia.miravia_table SET is_exported = %s WHERE lead_number = %s", (day,lead_number,))

    # Commit the changes and close the database connection
    conn.commit()
    conn.close()
    cursor.close()

    # Return a success message or any desired response
    return jsonify({'message': 'Update successful'})


# Replace the existing `/move_to_google_sheet` route with the updated code
@app.route('/move_to_google_sheet')
def move_to_google_sheet():
    import gspread
    import requests
    from google.auth import exceptions, impersonated_credentials
    from google.auth.transport.requests import Request
    from google.oauth2 import service_account
    # Load the credentials from the service account key file
    credentials = service_account.Credentials.from_service_account_file('miravia-credentials.json')

    # Check if the credentials have expired and refresh them if necessary
    if credentials.expired:
        try:
            credentials.refresh(Request())
        except exceptions.RefreshError as e:
            return f"Failed to refresh credentials: {e}", 500

    # Authorize using the impersonated credentials
    imp_credentials = impersonated_credentials.Credentials(
        source_credentials=credentials,
        target_principal='miravia-user@miravia-389703.iam.gserviceaccount.com',  # Replace with the email of the target account
        target_scopes=['https://www.googleapis.com/auth/spreadsheets']
    )

    try:
        # Authorize and open the Google Sheet
        gc = gspread.authorize(imp_credentials)
        sheet = gc.open_by_key('1DKa-U5fgqVKzp_OoqzRoDH2PE0AV38M6WBY5NX3Nn_8').sheet1

        # Fetch the data from the API
        response = requests.get('http://127.0.0.1:5000/api/data')
        data = response.json()
        print(data)

        # Iterate over the data and write it to the Google Sheet
        for item in data['data']:
            row = [item['assignment_date'], item['lead_number'], item['source'], item['url']]  # Add more fields as needed
            sheet.append_row(row)

        return "Data moved to Google Sheet successfully"
    except Exception as e:
        return f"Failed to move data to Google Sheet: {e}", 500


@app.route('/access')
def access():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()
    cursor.execute(f"SELECT email,role FROM miravia.miravia_users order by id asc")

    rows = cursor.fetchall()
    users = [row[0] for row in rows]
    roles = list(set([row[1] for row in rows]))
    
    cursor.close()
    conn.close()

    return render_template('access.html',users=users,roles=roles)

@app.route('/update_role', methods=['POST'])
def update_role():
    # Retrieve user and role values from the request
    user = request.form.get('user')
    role = request.form.get('role')

    # Perform the database update
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    try:
        cursor.execute("UPDATE miravia.miravia_users SET role = %s WHERE email = %s", (role, user))
        conn.commit()
        return jsonify({'message': 'Access granted successfully'})
    except Exception as e:
        conn.rollback()
        return jsonify({'message': 'An error occurred while granting access', 'error': str(e)})
    finally:
        conn.close()

@app.route('/create_user', methods=['GET', 'POST'])
def create_user():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()
    if request.method == 'POST':
        # Get form data
        username = request.form['username']
        role = request.form['role']
        password = request.form['password']

        cursor.execute("select * from miravia.miravia_users where email='%s'"%username)
        row = cursor.fetchone()

        if not row:
            try:
                # Define the SQL query
                query = "INSERT INTO miravia.miravia_users (email, password, role) VALUES (%s, %s, %s)"

                # Execute the query with the user data
                cursor.execute(query, (username, password, role))

                # Commit the changes to the database
                conn.commit()

                # Print a success message (you can modify this as needed)
                return jsonify({'message': 'User create successfully!'})
            except Exception as e:
                # Handle any errors that occur during the database operation
                conn.rollback()
                return jsonify({'message': 'Error creating user', 'error': str(e)})
                
            finally:
                # Close the database connection and cursor
                cursor.close()
                conn.close()
        else:
            # Username already exists
            return jsonify({'message': 'Username already exists!'})
        
@app.route('/delete_user', methods=['GET', 'POST'])
def delete_user():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()
    if request.method == 'POST':
        # Get form data
        username = request.form['username']

        cursor.execute("select * from miravia.miravia_users where email='%s'"%username)
        row = cursor.fetchone()
        if row:
            try:
                # Define the SQL query
                query = "delete from miravia.miravia_users where email= %s "

                # Execute the query with the user data
                cursor.execute(query, (username,))

                # Commit the changes to the database
                conn.commit()

                # Print a success message (you can modify this as needed)
                return jsonify({'message': 'User deleted!'})
            except Exception as e:
                # Handle any errors that occur during the database operation
                conn.rollback()
                return jsonify({'message': 'Error deleting user', 'error': str(e)})
                
            finally:
                # Close the database connection and cursor
                cursor.close()
                conn.close()
        else:
            # Username already exists
            return jsonify({'message': 'Username not found!'})

@app.route('/error-submit')
def error_submit():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    select_records_query = """SELECT id, lead_number, pic, error_type, error_date AS error_date FROM miravia.miravia_error_table order by id desc"""
    cursor.execute(select_records_query)
    errors = cursor.fetchall()

    return render_template('error.html', errors=errors)

@app.route('/save-error-data', methods=['POST'])
def save_error_data():
    lead_number = request.json.get('leadNumber')
    error_types = request.json.get('errorTypes')

    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    select_pic_query = """select pic from miravia.miravia_table where lead_number = %s"""
    cursor.execute(select_pic_query, (lead_number,))
    row = cursor.fetchone()
    if row:
        pic = row[0]
        
        insert_query = """
            INSERT INTO miravia.miravia_error_table (
                lead_number, error_type, pic, error_date
            )
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (lead_number) DO UPDATE SET
            error_type = EXCLUDED.error_type,
            pic = EXCLUDED.pic,
            error_date = EXCLUDED.error_date
        """

        try:
            cursor.execute(insert_query,(lead_number,' | '.join(error_types),pic,datetime.now().strftime("%d/%m/%Y")))
            conn.commit()
            return jsonify({'message':'success','lead_number':lead_number,'error_types':' | '.join(error_types),'pic':pic})
        except Exception as e:
            cursor.close()
            conn.close()
            raise e
    else:
        return jsonify({'message':'Lead Number not found!'})
    
@app.route('/delete-error-data', methods=['POST'])
def delete_error_data():
    lead_number = request.json.get('leadNumber')
    print(lead_number)

    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    select_pic_query = """select pic from miravia.miravia_error_table where lead_number = %s"""
    cursor.execute(select_pic_query, (lead_number,))
    row = cursor.fetchone()
    if row:
        pic = row[0]
        
        delete_query = """
            delete from miravia.miravia_error_table where lead_number = %s
        """

        try:
            cursor.execute(delete_query,(lead_number,))
            conn.commit()
            return jsonify({'message':'success','lead_number':lead_number,'pic':pic})
        except Exception as e:
            cursor.close()
            conn.close()
            raise e
    else:
        return jsonify({'message':'Lead Number not found!'})

@app.route('/revenue', methods=['GET', 'POST'])
def revenue():
    conn = psycopg2.connect(app.config['DATABASE_URI'])
    cursor = conn.cursor()

    current_year = datetime.now().year
    if request.method == 'POST':
        selected_month = request.form.get('month', '')
        selected_year = request.form.get('year', '') 
    else:
        selected_month = request.args.get('month', '')
        selected_year = request.args.get('year', '')

    exclusion_conditions = {
        "product_service": ["N/A", "Only Services"],
        "local_spanish_seller": ["N"],
        "feedback_final": ["MAINTENANCE WEBSITE"],
        "qualification_feedback": ["INACTIVE PAGE - Mainteinance"],
        "url": ["", "\s|@|amazon\.|Lazada|miravia.es|etsy|ebay\.|amzn.to"],
    }

    # Define the duration for each month
    month_duration = {
        1: (25, 21),
        2: (22, 21),
        3: (22, 21),
        4: (22, 21),
        5: (22, 21),
        6: (22, 21),
        7: (22, 21),
        8: (22, 21),
        9: (22, 21),
        10: (22, 23),
        11: (24, 23),
        12: (24, 22)
    }
    revenue_data = []
    paidlines_data = []
    error_count_data = []

    for month,duration in month_duration.items():
        start_day, end_day = duration
        if month == 1:
            start_date = f"{current_year}-12-{start_day:02d}"
            end_date = f"{current_year}-1-{end_day:02d}"
        else:
            month = month - 1
            start_date = f"{current_year}-{month:02d}-{start_day:02d}"
            end_date = f"{current_year}-{month+1:02d}-{end_day:02d}"

        # Total lines count for canvas
        select_revenue_query = f"""
            SELECT COUNT(*) FROM miravia.miravia_table
            WHERE TO_DATE(end_time, 'YYYY-mm-dd')::DATE >= DATE '{start_date}'::DATE
            AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE <= DATE '{end_date}'::DATE
        """
        cursor.execute(select_revenue_query)
        revenue_count = cursor.fetchone()[0]
        revenue_data.append(revenue_count)

        # Paid lines count for canvas
        exclusion_query_paid = ""
        for column, values in exclusion_conditions.items():
            if column == "url":
                value_conditions = [
                    f"url <> ''",
                    f"url ~ '\w+\.\w+' ",
                    f"url !~* '{'|'.join([x for x in values if x !=''])}'"
                ]
                exclusion_query_paid += f"AND ({' AND '.join(value_conditions)}) "
            else:
                value_conditions = [f"{column} <> '{value}'" for value in values]
                exclusion_query_paid += f"AND ({' AND '.join(value_conditions)}) "

        select_paid_lines_query = f"""
            SELECT COUNT(*) FROM miravia.miravia_table
            WHERE TO_DATE(end_time, 'YYYY-mm-dd')::DATE >= DATE '{start_date}'::DATE
            AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE <= DATE '{end_date}'::DATE {exclusion_query_paid}
        """
        
        cursor.execute(select_paid_lines_query)
        paid_count = cursor.fetchone()[0]
        paidlines_data.append(paid_count)
        #print(start_date,end_date,paid_count)

        # Total lines count for canvas
        select_error_query = f"""
            SELECT COUNT(*) FROM miravia.miravia_error_table
                WHERE TO_DATE(error_date, 'DD/MM/YYYY') >= DATE '{start_date}'::DATE
                AND TO_DATE(error_date, 'DD/MM/YYYY') <= DATE '{end_date}'::DATE
        """
        cursor.execute(select_error_query)
        error_count_all = cursor.fetchone()[0]
        error_count_data.append(error_count_all)

    select_pic_query = """SELECT email FROM miravia.miravia_users ORDER BY id ASC"""
    cursor.execute(select_pic_query)
    users = cursor.fetchall()

    rows = []
    for user in users:
        email = user[0]

        # Query miravia_table to get the count of non-empty 'pic' values for the current user
        select_total_lines_query = f"""
            SELECT COUNT(*) FROM miravia.miravia_table
            WHERE pic = '{email}'
        """
        cursor.execute(select_total_lines_query)
        total_lines = cursor.fetchone()[0]

        exclusion_query = ""
        for column, values in exclusion_conditions.items():
            if column == "url":
                value_conditions = [
                    f"url <> ''",
                    f"url ~ '\w+\.\w+' ",
                    f"url !~* '{'|'.join([x for x in values if x !=''])}'"
                ]
                exclusion_query += f"AND ({' AND '.join(value_conditions)}) "
            else:
                value_conditions = [f"{column} <> '{value}'" for value in values]
                exclusion_query += f"AND ({' AND '.join(value_conditions)}) "

        select_total_lines_query = f"""
                SELECT COUNT(*) FROM miravia.miravia_table
                WHERE pic = '{email}'
                AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE >= DATE '{current_year}-1-01'::DATE
                AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE < (DATE '{current_year}-12-01'::DATE + INTERVAL '1 month')
            """
        cursor.execute(select_total_lines_query)
        total_lines = cursor.fetchone()[0]

        cursor.execute(f"SELECT COUNT(*) "
                    f"FROM miravia.miravia_table "
                    f"WHERE pic = '{email}' "
                    f" AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE >= DATE '{current_year}-1-01'::DATE"
                    f" AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE < (DATE '{current_year}-12-01'::DATE + INTERVAL '1 month')"
                    f"{exclusion_query}")
        paid_lines = cursor.fetchone()[0]

        # Query miravia_error_table to get the count of errors for the current user
        select_error_count_query = f"""
            SELECT COUNT(*) FROM miravia.miravia_error_table
            WHERE pic = '{email}'
            AND TO_DATE(error_date, 'dd/mm/YYYY')::DATE >= DATE '{current_year}-1-01'::DATE
            AND TO_DATE(error_date, 'dd/mm/YYYY')::DATE < (DATE '{current_year}-12-01'::DATE + INTERVAL '1 month')
        """
        cursor.execute(select_error_count_query)
        error_count = cursor.fetchone()[0]

        # Calculate accuracy based on the provided formula
        accuracy = 100 - (error_count / paid_lines) * 100 if error_count != 0 else 100

        #revenue = paid_lines * (accuracy/100) * 3000 
        revenue = paid_lines * 3000
        
        formatted_revenue = int(revenue) > 0 and f"{int(revenue):,}".replace(".", ",") or 0

        if selected_month and selected_year:
            start_day, end_day = month_duration[int(selected_month)]
            if int(selected_month) == 1:
                start_date = f"{current_year}-12-{start_day:02d}"
                end_date = f"{current_year}-1-{end_day:02d}"
            else:
                start_date = f"{current_year}-{(int(selected_month) - 1):02d}-{start_day:02d}"
                end_date = f"{current_year}-{(int(selected_month) - 1)+1:02d}-{end_day:02d}"
            select_total_lines_query = f"""
                SELECT COUNT(*) FROM miravia.miravia_table
                WHERE pic = '{email}'
                AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE >= DATE '{start_date}'::DATE
                AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE <= DATE '{end_date}'::DATE
            """
            cursor.execute(select_total_lines_query)
            total_lines = cursor.fetchone()[0]

            cursor.execute(f"SELECT COUNT(*) "
                       f"FROM miravia.miravia_table "
                       f"WHERE pic = '{email}' "
                       f"AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE >= DATE '{start_date}'::DATE "
                       f"AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE <= DATE '{end_date}'::DATE "
                       f"{exclusion_query}")
            paid_lines = cursor.fetchone()[0]

            # Query miravia_error_table to get the count of errors for the current user
            select_error_count_query = f"""
                SELECT COUNT(*) FROM miravia.miravia_error_table
                WHERE pic = '{email}'
                AND TO_DATE(error_date, 'DD/MM/YYYY') >= DATE '{start_date}'::DATE
                AND TO_DATE(error_date, 'DD/MM/YYYY') <= DATE '{end_date}'::DATE
            """
            cursor.execute(select_error_count_query)
            error_count = cursor.fetchone()[0]

            # Calculate accuracy based on the provided formula
            accuracy = 100 - (error_count / paid_lines) * 100 if error_count != 0 else 100

            #revenue = paid_lines * (accuracy/100) * 3000 
            revenue = paid_lines * 3000
            
            formatted_revenue = int(revenue) > 0 and f"{int(revenue):,}".replace(".", ",") or 0
            
        elif selected_year:
            select_total_lines_query = f"""
                SELECT COUNT(*) FROM miravia.miravia_table
                WHERE pic = '{email}'
                AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE >= DATE '{selected_year}-1-01'::DATE
                AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE < (DATE '{selected_year}-12-01'::DATE + INTERVAL '1 month')
            """
            cursor.execute(select_total_lines_query)
            total_lines = cursor.fetchone()[0]

            cursor.execute(f"SELECT COUNT(*) "
                       f"FROM miravia.miravia_table "
                       f"WHERE pic = '{email}' "
                       f" AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE >= DATE '{selected_year}-1-01'::DATE"
                       f" AND TO_DATE(end_time, 'YYYY-mm-dd')::DATE < (DATE '{selected_year}-12-01'::DATE + INTERVAL '1 month')"
                       f"{exclusion_query}")
            paid_lines = cursor.fetchone()[0]

            # Query miravia_error_table to get the count of errors for the current user
            select_error_count_query = f"""
                SELECT COUNT(*) FROM miravia.miravia_error_table
                WHERE pic = '{email}'
                AND TO_DATE(error_date, 'dd/mm/YYYY')::DATE >= DATE '{selected_year}-1-01'::DATE
                AND TO_DATE(error_date, 'dd/mm/YYYY')::DATE < (DATE '{selected_year}-12-01'::DATE + INTERVAL '1 month')
            """
            cursor.execute(select_error_count_query)
            error_count = cursor.fetchone()[0]

            # Calculate accuracy based on the provided formula
            accuracy = 100 - (error_count / paid_lines) * 100 if error_count != 0 else 100

            #revenue = paid_lines * (accuracy/100) * 3000 
            revenue = paid_lines * 3000
            
            formatted_revenue = int(revenue) > 0 and f"{int(revenue):,}".replace(".", ",") or 0

        rows.append((email, total_lines, paid_lines, error_count, round(accuracy,1), formatted_revenue))

    rows = [
        (
            email,
            int(total_lines),
            int(paid_lines),
            int(error_count),
            '.' in str(accuracy) and round(float(accuracy), 1) or accuracy,
            int(str(formatted_revenue).replace(',', ''))
        )
        for email, total_lines, paid_lines, error_count, accuracy, formatted_revenue in rows
    ]

    # Calculate the column counts excluding the first column
    column_counts = [sum(column) for column in zip(*(row[1:] for row in rows))]

    rows = [
        (
            email,
            f"{total_lines:,}",  # Add comma as thousand separator
            f"{paid_lines:,}",  # Add comma as thousand separator
            error_count,
            accuracy,  # Format accuracy with one decimal place
            int(formatted_revenue) > 0 and f"{int(formatted_revenue):,}".replace(".", ",") or 0  # Remove commas in revenue
        )
        for email, total_lines, paid_lines, error_count, accuracy, formatted_revenue in rows
    ]

    column_counts = ['.' in str(x) and f"{float(x):,.1f}" or int(x) > 0 and f"{int(x):,}".replace(".", ",") or 0 for x in column_counts]

    cursor.close()
    conn.close()

    if request.method == 'POST':
        return jsonify({'rows':rows,'column_counts':column_counts})
    
    return render_template('revenue.html', rows=rows, revenue_data=revenue_data,paidlines_data=paidlines_data,error_count_data=error_count_data,column_counts=column_counts)




if __name__ == '__main__':
    app.run(debug=True)
